import openpyxl
import requests
#读取excel的测试用例
def real_data(workbookname,sheetname):
  wb=openpyxl.load_workbook(workbookname)
  sheet=wb[sheetname]  #获取sheet
  max_row=sheet.max_row   #获取sheet里最大行数
  list_1=[]   #定义一个空列表，来接收所有的测试数据
  for i in range(2,max_row+1): #取头不取尾，左闭右开
    dict_1=dict(
    id=sheet.cell(row=i,column=1).value,
    url = sheet.cell(row=i, column=5).value,
    data = sheet.cell(row=i, column=6).value,
    expect= sheet.cell(row=i, column=7).value)
    list_1.append(dict_1)#把所有的测试数据追加到列表
  return list_1
  #发送请求
def api_func(url, data):
  header = {'X-Lemonban-Media-Type': 'lemonban.v2',
              'Content-Type': 'application/json'}
  res = requests.post(url=url, json=data, headers=header)
  response=res.json()
  return response

#写入d断言结果
def write_rusult(workbookname,sheetname,row,column,final_result):
  wb=openpyxl.load_workbook(workbookname)
  sheet=wb[sheetname]
  sheet.cell(row=row,column=column).value=final_result
  wb.save('test_case_api.xlsx')
  #取出数据
def execute_func(workbookname,sheetname):
  cases=real_data(workbookname,sheetname)
  for case in cases:
    id=case.get('id') #取出id
    url=case.get('url')
    data=case.get('data')
    expect=case.get('expect')
    data=eval(data)
    expect=eval(expect)
    expect_msg=expect.get('msg')
    #从excel取出来的都是str,eval:运行被字符串包裹的python表达式
    result=api_func(url=url,data=data)
    real_msg=result.get('msg')
    print('执行结果为：{}'.format(real_msg))
    print('执行结果为：{}'.format( expect_msg))
    if expect_msg==real_msg:
      print('这条测试用例通过')
      final_res='pass'
    else:
      print('这条测试用例不通过')
      final_res='fail'
    print('*'*30)



    write_rusult(workbookname,sheetname,id+1,8,final_res)
execute_func('test_case_api.xlsx','register')